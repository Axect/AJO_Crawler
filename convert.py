import json
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.styles.alignment import Alignment
from datetime import datetime
import re

def parse_deadline(deadline_str):
    if not deadline_str:
        return None

    # 괄호와 공백 제거
    deadline_str = deadline_str.strip('() ')

    # 'deadline ' 문자열 제거
    deadline_str = deadline_str.replace('deadline ', '')

    # 'No deadline specified' 처리
    if deadline_str.lower() == 'no deadline specified':
        return None

    # 'offers accepted', 'accepting applications' 등의 특수 케이스 처리
    special_cases = ['offers accepted', 'accepting applications', 'filled', 'withdrawn']
    for case in special_cases:
        if case in deadline_str.lower():
            deadline_str = deadline_str.lower().replace(case, '').strip(', ')
            break

    # 날짜 추출을 위한 정규표현식
    date_pattern = r'(\d{4}/\d{1,2}/\d{1,2})'
    match = re.search(date_pattern, deadline_str)
    
    if match:
        date_str = match.group(1)
        try:
            return datetime.strptime(date_str, '%Y/%m/%d')
        except ValueError:
            print(f"Warning: Could not parse date: {date_str}")
            return None
    else:
        print(f"Warning: No valid date found in: {deadline_str}")
        return None

def json_to_excel(json_file, excel_file):
    # JSON 파일 읽기
    with open(json_file, 'r', encoding='utf-8') as file:
        data = json.load(file)

    # 새 워크북 생성
    wb = Workbook()
    ws = wb.active
    ws.title = "Physics Postdocs"

    # 헤더 작성
    headers = list(data[0].keys())
    for col, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col, value=header)
        ws.cell(row=1, column=col).font = Font(bold=True)

    # 데이터 작성
    for row, item in enumerate(data, start=2):
        for col, (key, value) in enumerate(item.items(), start=1):
            cell = ws.cell(row=row, column=col)
            if key == 'deadline':
                parsed_date = parse_deadline(value)
                if parsed_date:
                    cell.value = parsed_date
                    cell.number_format = 'YYYY-MM-DD'
                else:
                    cell.value = value  # 원본 문자열 유지
            elif key == 'application_materials' and isinstance(value, list):
                cell.value = '\n'.join(f'• {material}' for material in value)
                cell.alignment = Alignment(wrapText=True, vertical='top')
            else:
                cell.value = value

    # 열 너비 자동 조정
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)  # 최대 너비를 50으로 제한
        ws.column_dimensions[column].width = adjusted_width

    # Excel 파일 저장
    wb.save(excel_file)
    print(f"Excel 파일이 성공적으로 생성되었습니다: {excel_file}")

# 함수 실행
json_to_excel('data/physics_postdocs_updated.json', 'data/physics_postdocs_updated.xlsx')
